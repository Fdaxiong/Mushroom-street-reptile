# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
import time
from pprint import pprint
import pandas as pd
from openpyxl import Workbook

wb = Workbook()
ws = wb.active


class MogujiePipeline(object):
    def process_item(self, item, spider):
        # 保存数据excel
        self.sava_data(item)
        # pprint(item)
        return item

    def sava_data(self, item):
        """
        保存数据函数写入Excel
        """
        # 大分类
        # ws["A1"] = "theme_bazaar"
        # 大分类链接
        # ws["B1"] = "theme_bazaar_url"
        # 大分类抓包数据链接
        # ws["C1"] = "bazaar_details_url"
        # 中级分类
        # ws["D1"] = "recommend"
        # 小分类
        # ws["C1"] = "classify_title"
        # 小分类链接
        # ws["F1"] = "classify_url"
        # 小分类ID
        # ws["G1"] = "classify_fcid"
        # 小分类索引
        # ws["H1"] = "index"
        # 商品标题
        # ws["I1"] = "commodity_title"
        # 商品图片链接
        # ws["J1"] = "commodity_img"
        # 商品url链接
        # ws["K1"] = "commodity_link"
        # 商品原价
        # ws["L1"] = "commodity_orgPrice"
        # 商品现价
        # ws["M1"] = "commodity_price"
        ws["A1"] = "大分类"
        ws["B1"] = "中级分类"
        ws["C1"] = "小分类"
        ws["D1"] = "商品标题"
        ws["E1"] = "价格"
        ws["F1"] = "原价"
        ws["G1"] = "小分类链接"
        ws["H1"] = "商品图片链接"
        ws["I1"] = "商品链接"

        ws.append(
            [
                # item["theme_bazaar"],
                # item["theme_bazaar_url"],
                # item["bazaar_details_url"],
                # item["recommend"],
                # item["classify_title"],
                # item["classify_url"],
                # item["classify_fcid"],
                # item["index"],
                # item["commodity_title"],
                # item["commodity_img"],
                # item["commodity_link"],
                # item["commodity_orgPrice"],
                # item["commodity_price"],

                item["theme_bazaar"],
                item["recommend"],
                item["classify_title"],
                item["commodity_title"],
                item["commodity_price"],
                item["commodity_orgPrice"],
                item["classify_url"],
                item["commodity_img"],
                item["commodity_link"],

            ]
        )
        pprint(item)
        wb.save("蘑菇街_再也不修改版.xlsx")
        # wb.save("蘑菇街_完整版.xlsx")
